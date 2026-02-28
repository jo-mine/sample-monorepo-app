"""
指定シートのみを新しいExcelファイルに抽出するスクリプト
グラフ・図形・ピボットテーブルなどは、openpyxl が対応している範囲で可能な限り保持されます。
ただし、openpyxl で未対応または一部のみ対応している要素（特にピボットテーブル周辺の機能など）は、
保存時に失われる場合があります。

使用方法:
  python3 extract_sheet.py <入力xlsxパス> <シート名> <出力xlsxパス>
"""

import sys
import openpyxl


def extract_sheet(src: str, sheet_name: str, out_path: str) -> None:
    wb = openpyxl.load_workbook(src)

    if sheet_name not in wb.sheetnames:
        print(f"エラー: シート '{sheet_name}' が見つかりません", file=sys.stderr)
        print(f"利用可能なシート: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    # 対象シート以外を削除
    for name in list(wb.sheetnames):
        if name != sheet_name:
            del wb[name]

    wb.save(out_path)
    print(f"  シート '{sheet_name}' を抽出しました")


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print(
            "使用方法: python3 extract_sheet.py <入力xlsxパス> <シート名> <出力xlsxパス>",
            file=sys.stderr,
        )
        sys.exit(1)

    _, src, sheet_name, out_path = sys.argv
    extract_sheet(src, sheet_name, out_path)
